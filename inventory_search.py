# inventory_search.py
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.styles.colors import COLOR_INDEX

SHEET_NAME = "재고장"

# Column names in the source sheet
COL_CATEGORY = "분류"
COL_P = "P코드"
COL_T = "T코드"
COL_U = "U코드"
COL_ITEM = "품목코드"
COL_NAME = "품명"
COL_STOCK = "최종재고"
COL_POWER = "파워"
COL_COLOR = "컬러"
COL_COLOR_CODE = "컬러코드"
COL_TONE = "톤수"
COL_GROUP_PRODUCT = "구분_제품"
COL_GROUP_ERP = "구분_ERP"
COL_GROUP = COL_GROUP_PRODUCT
COL_LOTNO = "LOTNO"
COL_STERILE_NO_CANDIDATES = ["멸균No.", "멸균NO.", "멸균no.", "멸균No", "멸균NO", "멸균no"]
COL_CYL = "CYL"
COL_AXIS = "AXIS"
COL_ADD = "ADD"
COL_COLOR_HEX = "컬러_색상"
SUMMARY_SHEET = "SUMMARY"
SUMMARY_TORIC_SHEET = "SUMMARY (ToricMF)"

DEFAULT_SEARCH_COLS = [COL_P, COL_T, COL_U, COL_NAME, COL_ITEM, COL_COLOR, COL_COLOR_CODE, COL_TONE, COL_LOTNO]

CODE_PATTERN = re.compile(r"^[PTU]\d+$", re.IGNORECASE)
CODE_COLOR_PATTERN = re.compile(r"^([PTU]\d+)(.+)$", re.IGNORECASE)
TCODE_NUM_PATTERN = re.compile(r"(\d+)")

FILL_POS = PatternFill("solid", fgColor="FFF2CC")   # 옅은 노랑
FILL_NEG_0_10 = PatternFill("solid", fgColor="E6E6E6")  # 옅은 회색
FILL_NEG_10_20 = PatternFill("solid", fgColor="FBE5D6")  # 옅은 주황
THIN = Side(style="thin", color="D0D0D0")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_RED = Font(color="FF0000")


def _normalize_series(s: pd.Series) -> pd.Series:
    return s.fillna("").astype(str).str.strip()

def _compact_upper(s: pd.Series) -> pd.Series:
    return s.fillna("").astype(str).str.upper().str.replace(r"\s+", "", regex=True)


def _excel_color_to_hex(cell) -> str:
    fill = cell.fill
    if fill is None or fill.patternType is None or fill.fgColor is None:
        return ""
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        rgb = color.rgb
        if len(rgb) == 8:
            rgb = rgb[2:]
        return f"#{rgb}"
    if color.type == "indexed" and color.indexed is not None:
        try:
            rgb = COLOR_INDEX[color.indexed]
            if rgb:
                if len(rgb) == 8:
                    rgb = rgb[2:]
                return f"#{rgb}"
        except Exception:
            return ""
    return ""


def _fmt_power_value(v):
    try:
        fv = float(v)
    except Exception:
        return v
    sign = "-" if fv <= 0 else "+"
    return f"{sign}{abs(fv):05.2f}"


def _power_to_float(power) -> float | None:
    if power is None:
        return None
    try:
        return float(str(power).strip())
    except Exception:
        return None


def _power_fill(power) -> PatternFill | None:
    v = _power_to_float(power)
    if v is None:
        return None
    if v > 0:
        return FILL_POS
    if v >= -10.00:
        return FILL_NEG_0_10
    if v >= -20.00:
        return FILL_NEG_10_20
    return None


def _tcode_sort_key(code: str) -> tuple[int, str]:
    s = str(code or "").strip().upper()
    m = TCODE_NUM_PATTERN.search(s)
    if m:
        return (int(m.group(1)), s)
    return (10**9, s)


def _apply_row_style(ws, row: int, start_col: int, end_col: int, fill: PatternFill | None):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        if fill is not None:
            cell.fill = fill


def _apply_sheet_borders(ws, min_row: int = 1, min_col: int = 1):
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = BORDER_THIN


def _should_keep_power(power, total_qty: float, keep_neg_10_25_plus: bool, keep_pos_range: bool) -> bool:
    v = _power_to_float(power)
    if v is None:
        return True
    # -00.00 ~ -10.00 구간은 항상 유지
    if -10.00 <= v <= 0.00:
        return True
    # +파워 구간은 하나라도 재고가 있으면 전체 유지
    if v > 0 and keep_pos_range:
        return True
    # -10.25 이하 구간은 하나라도 재고가 있으면 전체 유지
    if v <= -10.25 and keep_neg_10_25_plus:
        return True
    # +파워 또는 -10.25 이하 구간은 재고가 있을 때만 표시
    try:
        return float(total_qty) > 0
    except Exception:
        return False


def _collect_powers(ws, start_row: int) -> list[str]:
    powers = []
    r = start_row
    while True:
        v = ws.cell(row=r, column=2).value
        if v is None or str(v).strip() == "":
            break
        if str(v).strip() in ("합계", "총합계"):
            break
        powers.append(v)
        r += 1

    # Ensure -00.25 exists between -00.00 and -00.50
    target = -0.25
    vals = []
    for p in powers:
        try:
            vals.append(float(str(p).strip()))
        except Exception:
            vals.append(None)

    if target not in [v for v in vals if v is not None]:
        insert_idx = None
        for i, v in enumerate(vals):
            if v is None:
                continue
            if v == 0:
                insert_idx = i + 1
                continue
            if v < 0 and v > target:
                insert_idx = i + 1
                continue
            if v <= target:
                insert_idx = i
                break
        if insert_idx is None:
            insert_idx = len(powers)
        powers.insert(insert_idx, _fmt_power_value(target))

    return [_fmt_power_value(p) for p in powers]


def _find_header_column(ws, header: str) -> int | None:
    for col_idx in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col_idx).value
        if v is not None and str(v).strip() == header:
            return col_idx
    return None


def _load_color_column_from_workbook(wb, data_len: int) -> list[str]:
    ws = wb[SHEET_NAME]
    color_col = _find_header_column(ws, COL_COLOR)
    if color_col is None:
        # Fallback to old 위치(컬러가 5열이던 시절)
        color_col = 5
    colors = []
    for row_idx in range(2, data_len + 2):
        cell = ws.cell(row=row_idx, column=color_col)
        colors.append(_excel_color_to_hex(cell))
    return colors


def _ensure_compat_columns(df: pd.DataFrame) -> pd.DataFrame:
    # 구버전 호환: 최종재고가 없으면 재고를 사용
    if COL_STOCK not in df.columns and "재고" in df.columns:
        df[COL_STOCK] = df["재고"]
    # 구버전 호환: 구분_제품이 없으면 구분을 사용
    if COL_GROUP_PRODUCT not in df.columns and "구분" in df.columns:
        df[COL_GROUP_PRODUCT] = df["구분"]
    return df


def load_inventory(path: Path, load_color: bool = False) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=SHEET_NAME)

    df = _ensure_compat_columns(df)

    # Normalize key columns to strings
    for col in [COL_CATEGORY, COL_P, COL_T, COL_U, COL_ITEM, COL_NAME, COL_COLOR, COL_COLOR_CODE, COL_TONE, COL_GROUP_PRODUCT, COL_GROUP_ERP, COL_LOTNO, COL_CYL, COL_AXIS, COL_ADD]:
        if col in df.columns:
            df[col] = _normalize_series(df[col])

    # Stock: NaN -> 0
    if COL_STOCK in df.columns:
        df[COL_STOCK] = pd.to_numeric(df[COL_STOCK], errors="coerce").fillna(0)

    # Load fill color from Excel column E (컬러) if requested
    if load_color:
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            if SHEET_NAME in wb.sheetnames:
                df[COL_COLOR_HEX] = _load_color_column_from_workbook(wb, len(df))
        except Exception:
            df[COL_COLOR_HEX] = ""

    return df


def load_inventory_from_bytes(data: bytes, load_color: bool = False) -> pd.DataFrame:
    from io import BytesIO

    bio = BytesIO(data)
    df = pd.read_excel(bio, sheet_name=SHEET_NAME)

    df = _ensure_compat_columns(df)

    for col in [COL_CATEGORY, COL_P, COL_T, COL_U, COL_ITEM, COL_NAME, COL_COLOR, COL_COLOR_CODE, COL_TONE, COL_GROUP_PRODUCT, COL_GROUP_ERP, COL_LOTNO, COL_CYL, COL_AXIS, COL_ADD]:
        if col in df.columns:
            df[col] = _normalize_series(df[col])

    if COL_STOCK in df.columns:
        df[COL_STOCK] = pd.to_numeric(df[COL_STOCK], errors="coerce").fillna(0)

    if load_color:
        try:
            bio.seek(0)
            wb = load_workbook(bio, data_only=True, read_only=True)
            if SHEET_NAME in wb.sheetnames:
                df[COL_COLOR_HEX] = _load_color_column_from_workbook(wb, len(df))
        except Exception:
            df[COL_COLOR_HEX] = ""

    return df


def _open_workbook_from_source(path: Path | None, data: bytes | None):
    from io import BytesIO

    if data is not None:
        bio = BytesIO(data)
        return load_workbook(bio, data_only=True, read_only=True)
    if path is None:
        raise ValueError("Either path or data must be provided")
    return load_workbook(path, data_only=True, read_only=True)


def _find_summary_column(ws, pcode: str, color: str, name: str) -> int | None:
    max_col = ws.max_column
    p_norm = str(pcode).strip().upper()
    c_norm = str(color).strip().upper() if color else ""
    n_norm = str(name).strip().upper() if name else ""

    # Match by P코드 + 컬러 (if provided)
    for c in range(3, max_col + 1):
        p = ws.cell(row=3, column=c).value
        if p is None:
            continue
        if str(p).strip().upper() != p_norm:
            continue
        if c_norm:
            col_val = ws.cell(row=4, column=c).value
            if str(col_val).strip().upper() != c_norm:
                continue
        return c

    # If 컬러가 없거나 매칭 실패했으면 P코드만으로 첫 컬럼 선택
    if p_norm:
        for c in range(3, max_col + 1):
            p = ws.cell(row=3, column=c).value
            if p is None:
                continue
            if str(p).strip().upper() == p_norm:
                return c

    # Fallback: match by 품명 on row 2
    if n_norm:
        for c in range(3, max_col + 1):
            v = ws.cell(row=2, column=c).value
            if v is None:
                continue
            if str(v).strip().upper() == n_norm:
                return c

    return None


def build_summary_export(
    pcode: str,
    color: str,
    name: str,
    tcode: str | None = None,
    color_code: str | None = None,
    tone: str | None = None,
    cyl: str | None = None,
    axis: str | None = None,
    add: str | None = None,
    source_path: Path | None = None,
    source_bytes: bytes | None = None,
    use_toric: bool = False,
    remove_powers: bool = False,
) -> tuple[bytes, bool]:
    wb = _open_workbook_from_source(source_path, source_bytes)
    sheet_name = SUMMARY_TORIC_SHEET if use_toric else SUMMARY_SHEET
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"{sheet_name} sheet not found")
    ws = wb[sheet_name]

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "EXPORT"

    if use_toric:
        out_ws["C2"] = name
        out_ws.merge_cells("C2:E2")
        out_ws["C3"] = "CODE"
        out_ws["D3"] = tcode or ""
        out_ws["E3"] = pcode
        out_ws["C4"] = "CYL"
        out_ws["D4"] = cyl or ""
        out_ws["E4"] = ""
        out_ws.merge_cells("D4:E4")
        out_ws["C5"] = "AXIS"
        out_ws["D5"] = axis or ""
        out_ws["E5"] = ""
        out_ws.merge_cells("D5:E5")
        out_ws["C6"] = "ADD"
        out_ws["D6"] = add or ""
        out_ws["E6"] = ""
        out_ws.merge_cells("D6:E6")
        out_ws["C7"] = color or ""
        out_ws["D7"] = color_code or ""
        out_ws["E7"] = tone or ""
        out_ws["B7"] = ws.cell(row=7, column=2).value or "파워"
        powers = _collect_powers(ws, start_row=8)
        row = 8
    else:
        out_ws["C2"] = name
        out_ws.merge_cells("C2:E2")
        out_ws["C3"] = "CODE"
        out_ws["D3"] = tcode or ""
        out_ws["E3"] = pcode
        out_ws["C4"] = color or ""
        out_ws["D4"] = color_code or ""
        out_ws["E4"] = tone or ""
        out_ws["B4"] = ws.cell(row=4, column=2).value or "파워"
        powers = _collect_powers(ws, start_row=5)
        row = 5

    if not remove_powers:
        for power in powers:
            out_ws.cell(row=row, column=2, value=power)
            qty_cell = out_ws.cell(row=row, column=3, value=0)
            qty_cell.number_format = "#,##0"
            out_ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
            _apply_row_style(out_ws, row, 2, 5, _power_fill(power))
            row += 1

    out_ws.cell(row=row, column=2, value="합계")
    total_cell = out_ws.cell(row=row, column=3, value=0)
    total_cell.number_format = "#,##0"
    out_ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    _apply_row_style(out_ws, row, 2, 5, None)

    _apply_sheet_borders(out_ws, min_row=1, min_col=2)
    from io import BytesIO

    bio = BytesIO()
    out_wb.save(bio)
    return bio.getvalue(), True


def build_summary_export_multi(
    rows: list[dict],
    source_path: Path | None = None,
    source_bytes: bytes | None = None,
    use_toric: bool = False,
    remove_powers: bool = False,
) -> tuple[bytes, int]:
    wb = _open_workbook_from_source(source_path, source_bytes)
    sheet_name = SUMMARY_TORIC_SHEET if use_toric else SUMMARY_SHEET
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"{sheet_name} sheet not found")
    ws = wb[sheet_name]

    if use_toric:
        powers = _collect_powers(ws, start_row=8)
        header_row = 7
        power_start_row = 8
    else:
        powers = _collect_powers(ws, start_row=5)
        header_row = 4
        power_start_row = 5

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "EXPORT"

    out_ws.cell(row=header_row, column=2, value=ws.cell(row=header_row, column=2).value or "파워")

    # Build product -> power qty map from search result rows
    product_keys = []
    power_map = {}
    info_map = {}
    power_totals = {}
    for row in rows:
        pcode = str(row.get(COL_P, "")).strip()
        tcode = str(row.get(COL_T, "")).strip()
        color = str(row.get(COL_COLOR, "")).strip()
        color_code = str(row.get(COL_COLOR_CODE, "")).strip()
        name = str(row.get(COL_NAME, "")).strip()
        tone = str(row.get(COL_TONE, "")).strip()
        cyl = str(row.get(COL_CYL, "")).strip()
        axis = str(row.get(COL_AXIS, "")).strip()
        add = str(row.get(COL_ADD, "")).strip()
        group = str(row.get(COL_GROUP, "")).strip()
        power = str(row.get(COL_POWER, "")).strip()
        qty = row.get(COL_STOCK, 0)
        if use_toric:
            key = (name, pcode, tcode, color, color_code, tone, cyl, axis, add)
        else:
            key = (name, pcode, tcode, color, color_code, tone)
        if key not in power_map:
            power_map[key] = {}
            product_keys.append(key)
            info_map[key] = {
                COL_CYL: cyl,
                COL_AXIS: axis,
                COL_ADD: add,
                COL_GROUP: group,
            }
        p_norm = power.strip().upper()
        try:
            qty_val = float(qty)
        except Exception:
            qty_val = 0.0
        power_map[key][p_norm] = power_map[key].get(p_norm, 0.0) + qty_val
        power_totals[p_norm] = power_totals.get(p_norm, 0.0) + qty_val

    has_neg_10_25_stock = any(
        (_power_to_float(p) is not None and _power_to_float(p) <= -10.25 and power_totals.get(str(p).strip().upper(), 0.0) > 0)
        for p in powers
    )
    has_pos_stock = any(
        (_power_to_float(p) is not None and _power_to_float(p) > 0 and power_totals.get(str(p).strip().upper(), 0.0) > 0)
        for p in powers
    )

    filtered_powers = []
    if not remove_powers:
        for power in powers:
            p_norm = str(power).strip().upper()
            total_qty = power_totals.get(p_norm, 0.0)
            if _should_keep_power(power, total_qty, has_neg_10_25_stock, has_pos_stock):
                filtered_powers.append(power)

    if use_toric:
        def _num_or_inf(v):
            try:
                return float(str(v).strip())
            except Exception:
                return float("inf")

        product_keys.sort(
            key=lambda k: (
                _num_or_inf(k[6]),  # CYL
                _num_or_inf(k[7]),  # AXIS
                _num_or_inf(k[8]),  # ADD
                _tcode_sort_key(k[2]),  # tcode
                k[0],                   # name
                k[1],                   # pcode
            )
        )
    else:
        product_keys.sort(
            key=lambda k: (
                _tcode_sort_key(k[2]),  # tcode
                k[0],                   # name
                k[1],                   # pcode
                k[3],                   # color
                k[4],                   # color_code
                k[5],                   # tone
            )
        )

    for idx, key in enumerate(product_keys):
        if use_toric:
            name, pcode, tcode, color, color_code, tone, cyl, axis, add = key
        else:
            name, pcode, tcode, color, color_code, tone = key

        out_col = 3 + (idx * 3)

        out_ws.cell(row=2, column=out_col, value=name)
        out_ws.merge_cells(start_row=2, start_column=out_col, end_row=2, end_column=out_col + 2)
        out_ws.cell(row=3, column=out_col, value="CODE")
        out_ws.cell(row=3, column=out_col + 1, value=tcode)
        out_ws.cell(row=3, column=out_col + 2, value=pcode)
        if use_toric:
            info = info_map.get(key, {})
            group_val = str(info.get(COL_GROUP, "")).strip().upper()
            show_cyl_axis = True
            show_add = True
            if group_val == "TORIC":
                show_add = False
            elif group_val == "M/F":
                show_cyl_axis = False
            elif group_val == "TORIC+M/F":
                show_cyl_axis = True
                show_add = True

            out_ws.cell(row=4, column=out_col, value="CYL")
            out_ws.cell(row=4, column=out_col + 1, value=info.get(COL_CYL, "") if show_cyl_axis else "")
            out_ws.cell(row=4, column=out_col + 2, value="")
            out_ws.merge_cells(start_row=4, start_column=out_col + 1, end_row=4, end_column=out_col + 2)
            out_ws.cell(row=5, column=out_col, value="AXIS")
            out_ws.cell(row=5, column=out_col + 1, value=info.get(COL_AXIS, "") if show_cyl_axis else "")
            out_ws.cell(row=5, column=out_col + 2, value="")
            out_ws.merge_cells(start_row=5, start_column=out_col + 1, end_row=5, end_column=out_col + 2)
            out_ws.cell(row=6, column=out_col, value="ADD")
            out_ws.cell(row=6, column=out_col + 1, value=info.get(COL_ADD, "") if show_add else "")
            out_ws.cell(row=6, column=out_col + 2, value="")
            out_ws.merge_cells(start_row=6, start_column=out_col + 1, end_row=6, end_column=out_col + 2)
            out_ws.cell(row=7, column=out_col, value=color or "")
            out_ws.cell(row=7, column=out_col + 1, value=color_code or "")
            out_ws.cell(row=7, column=out_col + 2, value=tone or "")
        else:
            out_ws.cell(row=4, column=out_col, value=color or "")
            out_ws.cell(row=4, column=out_col + 1, value=color_code or "")
            out_ws.cell(row=4, column=out_col + 2, value=tone or "")

        total = 0.0
        local_map = power_map.get(key, {})
        if not remove_powers:
            for i, power in enumerate(filtered_powers):
                row = power_start_row + i
                p_norm = str(power).strip().upper()
                qty = local_map.get(p_norm, 0.0)
                out_ws.cell(row=row, column=2, value=power)
                qty_cell = out_ws.cell(row=row, column=out_col, value=qty)
                qty_cell.number_format = "#,##0"
                out_ws.merge_cells(start_row=row, start_column=out_col, end_row=row, end_column=out_col + 2)
                _apply_row_style(out_ws, row, 2, 2, _power_fill(power))
                _apply_row_style(out_ws, row, out_col, out_col + 2, _power_fill(power))
                try:
                    total += float(qty)
                except Exception:
                    pass
        else:
            try:
                total = float(sum(local_map.values()))
            except Exception:
                total = 0.0

        total_row = power_start_row + (len(filtered_powers) if not remove_powers else 0)
        out_ws.cell(row=total_row, column=2, value="합계")
        total_cell = out_ws.cell(row=total_row, column=out_col, value=total)
        total_cell.number_format = "#,##0"
        out_ws.merge_cells(
            start_row=total_row,
            start_column=out_col,
            end_row=total_row,
            end_column=out_col + 2,
        )
        _apply_row_style(out_ws, total_row, 2, 2, None)
        _apply_row_style(out_ws, total_row, out_col, out_col + 2, None)

    if not remove_powers:
        # 행 전체 합이 0일 때만 빨간색 표시
        for i, power in enumerate(filtered_powers):
            row = power_start_row + i
            row_total = 0.0
            p_norm = str(power).strip().upper()
            for key in product_keys:
                local_map = power_map.get(key, {})
                row_total += float(local_map.get(p_norm, 0.0) or 0.0)
            if row_total == 0:
                out_ws.cell(row=row, column=2).font = FONT_RED
                for idx in range(len(product_keys)):
                    out_col = 3 + (idx * 3)
                    out_ws.cell(row=row, column=out_col).font = FONT_RED

    _apply_sheet_borders(out_ws, min_row=1, min_col=2)
    from io import BytesIO

    bio = BytesIO()
    out_wb.save(bio)
    return bio.getvalue(), 0


def _term_mask(df: pd.DataFrame, term: str) -> pd.Series:
    term = term.strip()
    if not term:
        return pd.Series(False, index=df.index)

    term_upper = term.upper()
    code_color_match = CODE_COLOR_PATTERN.match(term_upper)
    is_code_like = bool(CODE_PATTERN.match(term_upper))

    masks = []

    if code_color_match and not is_code_like:
        code = code_color_match.group(1)
        color_term = code_color_match.group(2).strip()
        tone_term = ""
        tone_match = re.match(r"^(.*?)(\d+)$", color_term)
        if tone_match:
            color_term = tone_match.group(1).strip()
            tone_term = tone_match.group(2).strip()

        if COL_P in df.columns:
            masks.append(df[COL_P].str.upper() == code)
        if COL_T in df.columns:
            masks.append(df[COL_T].str.upper() == code)
        if COL_U in df.columns:
            masks.append(df[COL_U].str.upper() == code)

        color_masks = []
        if COL_COLOR in df.columns:
            term_compact = re.sub(r"\s+", "", color_term).upper()
            color_masks.append(_compact_upper(df[COL_COLOR]).str.contains(term_compact, na=False))
        if "컬러코드" in df.columns:
            term_compact = re.sub(r"\s+", "", color_term).upper()
            color_masks.append(_compact_upper(df["컬러코드"]).str.contains(term_compact, na=False))

        if tone_term and COL_TONE in df.columns:
            tone_mask = df[COL_TONE].str.upper() == tone_term.upper()
        else:
            tone_mask = None

        if color_masks:
            color_mask = color_masks[0]
            for m in color_masks[1:]:
                color_mask = color_mask | m

            if masks:
                code_mask = masks[0]
                for m in masks[1:]:
                    code_mask = code_mask | m
                combined = code_mask & color_mask
                if tone_mask is not None:
                    combined = combined & tone_mask
                return combined
            combined = color_mask
            if tone_mask is not None:
                combined = combined & tone_mask
            return combined

    if is_code_like:
        # Exact match for code fields
        if COL_P in df.columns:
            masks.append(df[COL_P].str.upper() == term_upper)
        if COL_T in df.columns:
            masks.append(df[COL_T].str.upper() == term_upper)
        if COL_U in df.columns:
            masks.append(df[COL_U].str.upper() == term_upper)
        # Allow partial match on item code for convenience
        if COL_ITEM in df.columns:
            masks.append(df[COL_ITEM].str.upper().str.contains(term_upper, na=False))
    else:
        # Partial match across name + codes
        for col in DEFAULT_SEARCH_COLS:
            if col in df.columns:
                masks.append(df[col].str.upper().str.contains(term_upper, na=False))

    if not masks:
        return pd.Series(False, index=df.index)

    mask = masks[0]
    for m in masks[1:]:
        mask = mask | m
    return mask


def _summarize_inventory(df: pd.DataFrame) -> pd.DataFrame:
    # Group by P/T + color + tone + power and keep a representative name
    group_cols = []
    if COL_CATEGORY in df.columns:
        group_cols.append(COL_CATEGORY)
    if COL_P in df.columns:
        group_cols.append(COL_P)
    if COL_T in df.columns:
        group_cols.append(COL_T)
    if COL_COLOR in df.columns:
        group_cols.append(COL_COLOR)
    if COL_COLOR_CODE in df.columns:
        group_cols.append(COL_COLOR_CODE)
    if COL_TONE in df.columns:
        group_cols.append(COL_TONE)
    if COL_POWER in df.columns:
        group_cols.append(COL_POWER)
    if COL_CYL in df.columns:
        group_cols.append(COL_CYL)
    if COL_AXIS in df.columns:
        group_cols.append(COL_AXIS)
    if COL_ADD in df.columns:
        group_cols.append(COL_ADD)

    if not group_cols:
        return df.head(0)

    result = (
        df.groupby(group_cols, dropna=False)
        .agg(
            **{
                COL_NAME: (COL_NAME, "first") if COL_NAME in df.columns else (COL_P, "first"),
                COL_STOCK: (COL_STOCK, "sum"),
                "LOTNO 합계수량": (COL_LOTNO, lambda s: s.notna().sum()) if COL_LOTNO in df.columns else (COL_STOCK, "count"),
                COL_COLOR_HEX: (COL_COLOR_HEX, "first") if COL_COLOR_HEX in df.columns else (COL_P, "first"),
            }
        )
        .reset_index()
        .sort_values(COL_STOCK, ascending=False)
    )

    if COL_STOCK in result.columns:
        result[COL_STOCK] = pd.to_numeric(result[COL_STOCK], errors="coerce").fillna(0).round(0).astype(int)

    # Power formatting: -00.00, -05.25, -04.75
    if COL_POWER in result.columns:
        result[COL_POWER] = result[COL_POWER].apply(_fmt_power_value)

    # Column order: P, T, 컬러, 톤수, 파워, 품명, 재고, LOTNO 합계수량
    preferred = [
        COL_CATEGORY,
        COL_P,
        COL_T,
        COL_COLOR,
        COL_COLOR_CODE,
        COL_TONE,
        COL_POWER,
        COL_CYL,
        COL_AXIS,
        COL_ADD,
        COL_NAME,
        COL_STOCK,
        "LOTNO 합계수량",
        COL_COLOR_HEX,
    ]
    cols = [c for c in preferred if c in result.columns]
    if cols:
        result = result[cols]

    return result


def summarize_inventory(df: pd.DataFrame) -> pd.DataFrame:
    return _summarize_inventory(df)


def search_inventory(df: pd.DataFrame, query: str) -> pd.DataFrame:
    # Split by spaces or commas
    terms = [t for t in re.split(r"[\s,]+", query) if t.strip()]
    if not terms:
        return df.head(0)

    mask = pd.Series(False, index=df.index)
    for term in terms:
        mask = mask | _term_mask(df, term)

    filtered = df[mask]
    return _summarize_inventory(filtered)


def lot_breakdown(df: pd.DataFrame, row: dict) -> pd.DataFrame:
    # Filter original rows to match selected summary row
    filters = []
    if COL_P in df.columns and row.get(COL_P) is not None:
        filters.append(df[COL_P] == str(row.get(COL_P, "")).strip())
    if COL_T in df.columns and row.get(COL_T) is not None:
        filters.append(df[COL_T] == str(row.get(COL_T, "")).strip())
    if COL_COLOR in df.columns and row.get(COL_COLOR) is not None:
        filters.append(df[COL_COLOR] == str(row.get(COL_COLOR, "")).strip())
    if COL_TONE in df.columns and row.get(COL_TONE) is not None:
        filters.append(df[COL_TONE] == str(row.get(COL_TONE, "")).strip())
    if COL_POWER in df.columns and row.get(COL_POWER) is not None:
        try:
            power_val = float(str(row.get(COL_POWER, "")).strip())
            power_series = pd.to_numeric(df[COL_POWER], errors="coerce").round(2)
            filters.append(power_series == round(power_val, 2))
        except Exception:
            filters.append(df[COL_POWER].astype(str) == str(row.get(COL_POWER, "")).strip())
    if COL_COLOR_CODE in df.columns and row.get(COL_COLOR_CODE) is not None:
        filters.append(df[COL_COLOR_CODE].astype(str) == str(row.get(COL_COLOR_CODE, "")).strip())
    if COL_CYL in df.columns and row.get(COL_CYL) is not None:
        filters.append(df[COL_CYL].astype(str) == str(row.get(COL_CYL, "")).strip())
    if COL_AXIS in df.columns and row.get(COL_AXIS) is not None:
        filters.append(df[COL_AXIS].astype(str) == str(row.get(COL_AXIS, "")).strip())
    if COL_ADD in df.columns and row.get(COL_ADD) is not None:
        filters.append(df[COL_ADD].astype(str) == str(row.get(COL_ADD, "")).strip())

    if not filters:
        return df.head(0)

    mask = filters[0]
    for f in filters[1:]:
        mask = mask & f

    filtered = df[mask]
    if COL_LOTNO not in filtered.columns:
        return filtered.head(0)

    sterile_col = None
    for c in COL_STERILE_NO_CANDIDATES:
        if c in filtered.columns:
            sterile_col = c
            break

    if sterile_col is not None:
        lot_df = (
            filtered.groupby(COL_LOTNO, dropna=False)
            .agg(
                **{
                    "멸균No.": (
                        sterile_col,
                        lambda s: ", ".join(
                            sorted(
                                {
                                    str(v).strip()
                                    for v in s.fillna("")
                                    if str(v).strip()
                                }
                            )
                        ),
                    ),
                    COL_STOCK: (COL_STOCK, "sum"),
                }
            )
            .reset_index()
            .sort_values(COL_STOCK, ascending=False)
        )
    else:
        lot_df = (
            filtered.groupby(COL_LOTNO, dropna=False)[COL_STOCK]
            .sum()
            .reset_index()
            .sort_values(COL_STOCK, ascending=False)
        )
    if COL_STOCK in lot_df.columns:
        lot_df[COL_STOCK] = pd.to_numeric(lot_df[COL_STOCK], errors="coerce").fillna(0).round(0).astype(int)

    return lot_df


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python inventory_search.py <query> [excel_path]")
        return 1

    query = sys.argv[1]
    path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(__file__).parent / "재고관련 프로그램제작.xlsx"

    if not path.exists():
        print(f"File not found: {path}")
        return 1

    df = load_inventory(path)
    result = search_inventory(df, query)

    if result.empty:
        print("No matches")
        return 0

    # Print a readable table
    print(result.to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
