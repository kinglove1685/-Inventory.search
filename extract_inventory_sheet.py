from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook


DEFAULT_SHEETS = ("재고장", "멸균넘버 기준표")


def extract_sheet_values(src: Path, dst: Path, sheet_names: tuple[str, ...] = DEFAULT_SHEETS) -> None:
    wb = load_workbook(src, data_only=True, read_only=True)
    missing = [name for name in sheet_names if name not in wb.sheetnames]
    if missing:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Missing sheets: {', '.join(missing)}. Available sheets: {available}")

    out_wb = Workbook()
    first_sheet = True
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        if first_sheet:
            out_ws = out_wb.active
            out_ws.title = sheet_name
            first_sheet = False
        else:
            out_ws = out_wb.create_sheet(title=sheet_name)

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                out_ws.cell(row=row_idx, column=col_idx, value="" if value is None else str(value).strip())

    out_wb.save(dst)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Create a lightweight workbook by copying values from selected sheets."
    )
    parser.add_argument("source", type=Path, help="Source Excel file path")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("inventory_dummy.xlsx"),
        help="Output Excel file path (default: inventory_dummy.xlsx)",
    )
    parser.add_argument(
        "-s",
        "--sheet",
        dest="sheets",
        action="append",
        help="Sheet name to extract. Repeat this option to include multiple sheets.",
    )
    args = parser.parse_args()

    selected_sheets = tuple(args.sheets) if args.sheets else DEFAULT_SHEETS
    extract_sheet_values(args.source, args.output, selected_sheets)
    print(f"Done: {args.output}")


if __name__ == "__main__":
    main()
